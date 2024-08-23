import os

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font

from utils import fill_cell, clean_up, fill_cell_link

FONT_SIZE = 16
COLUMN_SUM = 12
BACK_SLASH = "\\"


class Team:
    def __init__(self, name, race_name, race_partition, work_name):
        self.name = name
        self.race_name = race_name
        self.race_partition = race_partition
        self.work_name = work_name

    def __str__(self):
        return str(self.name) + ' ' + str(self.race_name) + ' ' + str(self.race_partition) + ' ' + str(self.work_name)


race_array = ['工业制造',
              '现代农业',
              '商贸流通',
              '交通运输',
              '金融服务',
              '科技创新',
              '文化旅游',
              '医疗健康',
              '应急管理',
              '气象服务',
              '城市治理',
              '绿色低碳']


def scdex(input_namelist, file_base):
    race_dict = read_namelist_to_team(input_namelist)
    table_prefix = '2024年“数据要素×”大赛四川分赛初赛评分表_'
    table_subfix = '.xlsx'
    for partition in race_array:
        dest_excel = file_base + BACK_SLASH + partition + BACK_SLASH + table_prefix + partition + table_subfix
        dest_dir = file_base + BACK_SLASH + partition
        if not os.path.isdir(dest_dir) or not os.path.exists(dest_excel):
            continue
        write_to_score_sheet(race_dict, dest_excel, dest_dir)


def write_to_score_sheet(race_dict, workbook_path, file_base):
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb['Sheet1']

    clean_up(sheet, 5, 10)

    for race_key in race_dict:
        race_partition_dict = race_dict[race_key]
        row = 5
        for key in race_partition_dict:
            partition_array = race_partition_dict[key]
            partition_base = file_base + BACK_SLASH + key
            for i in range(len(partition_array)):
                team: Team = partition_array[i]
                if team.work_name is None or team.work_name == '':
                    continue

                team_prefix = '【'
                team_subfix = '】'
                team_base = partition_base + BACK_SLASH + team_prefix + team.name + team_subfix
                # program_base = team_base + BACK_SLASH + '项目申报书'

                ## tmp logic
                if not os.path.isdir(team_base):
                    continue

                fill_cell(sheet.cell(row, column=1), str(row - 4))
                fill_cell(sheet.cell(row, column=2), str(team.race_name))
                fill_cell(sheet.cell(row, column=3), str(team.race_partition))
                fill_cell(sheet.cell(row, column=4), str(team.name))
                link_file = ".\\" + key + BACK_SLASH + team_prefix + team.name + team_subfix + BACK_SLASH + get_program_file_name(team_base)
                fill_cell_link(sheet.cell(row, column=5), str(team.work_name), link_file)

                # 累加和
                cell = sheet.cell(row, column=10)
                cell.value = "=SUM({}:{})".format("F" + str(row), "I" + str(row))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                font = Font('Times New Roman', color='000000', bold=False, size=FONT_SIZE)
                cell.font = font

                sheet.row_dimensions[row].height = 100
                row += 1

    wb.save(workbook_path)


def get_program_file_name(work_dir) -> str:
    for file in os.listdir(work_dir):
        tail = os.path.splitext(file)[-1][1:]
        if file.__contains__("申报书") and tail == "pdf":
            return file

    print("Error in find program file: " + str(work_dir))
    return ''


def read_namelist_to_team(workbook_path) -> dict:
    # sheet
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb['参赛团队列表824']
    team_set = set()
    race_dict = {}

    for row in range(2, sheet.max_row):
        # 参赛团队
        team_name = sheet.cell(row, column=3).value
        if team_name in team_set:
            continue

        team_set.add(team_name)

        # 赛道名称
        race_name = sheet.cell(row, column=1).value
        # 参赛方向
        race_partition = sheet.cell(row, column=2).value

        # 作品名称
        work_name = sheet.cell(row, column=7).value

        team = Team(team_name, race_name, race_partition, work_name)

        if race_dict.get(race_name) is None:
            race_dict[race_name] = {}

        if not race_dict[race_name].__contains__(race_partition):
            race_dict[race_name][race_partition] = []

        race_dict[race_name][race_partition].append(team)

    print(str(race_dict))
    return race_dict
