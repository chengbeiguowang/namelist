import os

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font,Border, Side

PROVINCE_NUM = 32
ENTERPRISE_NUM = 36
COLUMN_SUM = 15


def handle_law_namelist(race_base, workbook_path):
    wb = openpyxl.load_workbook(workbook_path)

    sheet_name = ['省直部门+市州（共32份）', '企业+区县（共36份）']
    cleanup()

    sheet = wb[sheet_name[0]]
    handle_namelist_inner(sheet, race_base + "\\省直部门＋市州", "省直部门＋市州", PROVINCE_NUM)
    sheet = wb[sheet_name[1]]
    handle_namelist_inner(sheet, race_base + "\\企业＋区县", "企业＋区县", ENTERPRISE_NUM)

    wb.save(workbook_path)


def handle_namelist_inner(sheet, race_path, dir_name, item_num):
    files = os.listdir(race_path)
    # 对数字进行排序
    files.sort(key=lambda x: int(str(x).split('.')[0]))

    _sheet_offset = 5
    for i in range(0, item_num):
        row = i + _sheet_offset
        # index and name check
        sheet_index = str(sheet.cell(row, 1).value)
        sheet_name = sheet.cell(row, 2).value.strip()
        array = files[i].split('.')
        file_index = str(array[0])
        file_name = array[1]
        if sheet_index != file_index or sheet_name != file_name:
            raise Exception("sheet_name:" + sheet_name + " " + "file_name:" + file_name)

        work_index = 4
        fill_cell_link(sheet.cell(row, work_index), dir_name + "\\" + files[i])
        print(str(sheet.cell(row, 1).value) + " " + files[i])

        cell = sheet.cell(row, column=COLUMN_SUM)
        cell.value = "=SUM({}:{})".format("E" + str(row), "N" + str(row))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

def fill_cell_link(cell, link):
    cell.hyperlink = link
    # cell.style = "Hyperlink"
    font = Font('仿宋', color='0563C1', bold=False, size=12)
    cell.font = font
    # cell.style = "Hyperlink"
    # cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
    #border = Border(left=Side(style="thin", color='000000'), right=Side(style="thin",color='000000'), top=Side(style="thin", color='000000'), bottom=Side(style="thin",  color='000000'))
    #cell.border = border

def cleanup():
    pass
