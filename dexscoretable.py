import os.path

import openpyxl

from math_utils import is_float
from namelist import clean_up
from utils import fill_cell, fill_cell_if_empty

BACK_SLASH = "\\"

race_array = ['工业制造',
              '现代农业',
              '商贸流通',
              '交通运输',
              '金融服务',
              '科技创新',
              '文化旅游',
              '医疗健康',
              '应急管理',
              '气象服务',
              '城市治理',
              '绿色低碳']


def compute_dex_score(root_dir):
    summary_table = root_dir + BACK_SLASH + '汇总表.xlsx'
    for race in race_array:
        score_table_base = root_dir + BACK_SLASH + race
        if os.path.isdir(score_table_base):
            compute_race_score(race, score_table_base, summary_table)


def compute_race_score(race_name, race_dir, summary_table):
    count = 0
    total_dict = dict()
    print(race_name)
    for score_table in os.listdir(race_dir):
        table_dict = read_score_table(race_dir + BACK_SLASH + str(score_table))
        total_dict.update(table_dict)
        count += 1

    if count != 5:
        print("error in score table num:" + str(count) + " path:" + race_dir)

    write_score_to_summary_table(total_dict, summary_table, race_name)


def write_score_to_summary_table(total_dict, summary_table, race_name):
    wb = openpyxl.load_workbook(summary_table)
    sheet = wb[race_name]
    clean_up(sheet, 1, 12)

    ### write header
    fill_cell(sheet.cell(1, column=1), '序号')
    fill_cell(sheet.cell(1, column=2), '赛道名称')
    fill_cell(sheet.cell(1, column=3), '赛题名称')
    fill_cell(sheet.cell(1, column=4), '团队名称')
    fill_cell(sheet.cell(1, column=5), '团队作品')

    score_column_start = column = 6
    size = 0

    for expert_name in total_dict:
        # 写入专家姓名
        fill_cell(sheet.cell(1, column=column), expert_name)
        score_array = total_dict[expert_name]
        row = 2
        index = 0
        size = len(score_array)
        for item in score_array:
            fill_cell_if_empty(sheet.cell(row + index, column=1), item['num'])
            fill_cell_if_empty(sheet.cell(row + index, column=2), item['race_name'])
            fill_cell_if_empty(sheet.cell(row + index, column=3), item['race_partition'])
            fill_cell_if_empty(sheet.cell(row + index, column=4), item['team_name'])
            fill_cell_if_empty(sheet.cell(row + index, column=5), item['work_name'])
            cell = sheet.cell(row + index, column=5)
            if cell.value is None:
                fill_cell(cell, item['work_name'])
            else:
                if cell.value != item['work_name']:
                    print(cell.value + " " + item['work_name'] + " do not match")

            fill_cell(sheet.cell(row + index, column=column), float(item['score']))
            index += 1

        column += 1

    avg_column = column
    fill_cell(sheet.cell(1, column=avg_column), '平均分')
    without_ends_avg_column = column + 1
    #fill_cell(sheet.cell(1, column=without_ends_avg_column), '掐头去尾平均')

    for i in range(2, 1 + size + 1):
        sum_value = float(0)
        float_array = []
        for j in range(score_column_start, avg_column):
            float_array.append(sheet.cell(i, j).value)
            sum_value += float(sheet.cell(i, j).value)

        # 计算平均数
        avg = sum_value / (avg_column - score_column_start)
        fill_cell(sheet.cell(i, column=avg_column), str(avg))

        # 计算掐头去尾平均数
        #float_array.sort()
        #without_ends_avg = round((float_array[1] + float_array[2] + float_array[3]) / 3, 2)
        #fill_cell(sheet.cell(i, column=without_ends_avg_column), str(without_ends_avg))

    wb.save(summary_table)


def read_score_table(score_table_path) -> dict:
    wb = openpyxl.load_workbook(score_table_path, data_only=True)
    sheet = wb['Sheet1']

    table_dict = {}
    count = 0

    expert_name = sheet.cell(2, column=1).value.split(':')[1].strip()

    team_array = []

    row = 5
    cell = sheet.cell(row, column=1)
    while cell.value is not None:
        # 序号
        num = sheet.cell(row, column=1).value
        # 赛道名称
        race_name = sheet.cell(row, column=2).value
        # 赛题名称
        race_partition = sheet.cell(row, column=3).value.strip()
        # 团队名称
        team_name = sheet.cell(row, column=4).value.strip()
        # 团队作品
        work_name = sheet.cell(row, column=5).value.strip()

        # 校验
        if not validate(sheet.cell(row, column=6).value, 0, 30):
            print(expert_name + " " + race_name + " " + team_name + " " + work_name + " ")
        if not validate(sheet.cell(row, column=7).value, 0, 30):
            print(expert_name + " " + race_name + " " + team_name + " " + work_name + " ")
        if not validate(sheet.cell(row, column=8).value, 0, 30):
            print(expert_name + " " + race_name + " " + team_name + " " + work_name + " ")
        if not validate(sheet.cell(row, column=9).value, 0, 10):
            print(expert_name + " " + race_name + " " + team_name + " " + work_name + " ")

        # 评分合计
        score = str(sheet.cell(row, column=10).value)
        if not validate(score, 0, 100):
            print(expert_name + " " + race_name + " " + team_name + " " + work_name + " " + score)

        team_item = dict()
        team_item['num'] = num
        team_item['race_name'] = race_name
        team_item['race_partition'] = race_partition
        team_item['team_name'] = team_name
        team_item['work_name'] = work_name
        team_item['score'] = score
        if not (score.isdigit() or is_float(score)):
            print("not number:" + str(team_item))

        team_array.append(team_item)
        count += 1

        row += 1
        cell = sheet.cell(row, column=1)

    table_dict[expert_name] = team_array
    print("专家：" + str(expert_name) + " item count:" + str(count))

    return table_dict


def validate(value, lower, upper):
    if value is None or value == '':
        return True

    if type(value) == float:
        return lower <= value <= upper

    if type(value) == int:
        return lower <= value <= upper

    if type(value) == str:
        if is_float(value):
            return lower <= float(value) <= upper
        else:
            return lower <= int(value) <= upper

    return False
