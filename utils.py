from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


def fill_cell_if_empty(cell, content, font_size=16):
    if cell.value is None:
        fill_cell(cell, content, font_size)


def fill_cell(cell, content, font_size=16):
    cell.value = content
    cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
    font = Font('宋体', color='000000', bold=False, size=font_size)
    cell.font = font


def fill_cell_link(cell, file_name, link, font_size=16):
    cell.value = file_name
    cell.hyperlink = link
    font = Font('宋体', color='0563C1', bold=False, size=font_size)
    cell.font = font
    # cell.style = "Hyperlink"
    cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)


def is_empty(value) -> bool:
    return value is None or value == ''


def clean_up(sheet: Worksheet, row_start, column_max, font_size=16):
    _row_start = row_start
    cell = sheet.cell(_row_start, 1)
    while not is_empty(cell.value):
        for i in range(1, column_max + 1):
            cell = sheet.cell(_row_start, i)
            cell.value = None
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            font = Font('Times New Roman', color='000000', bold=False, size=font_size)
            cell.font = font
        _row_start += 1
        cell = sheet.cell(_row_start, 1)
