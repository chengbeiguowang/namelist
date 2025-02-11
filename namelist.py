import os

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font

FONT_SIZE = 14
COLUMN_SUM = 12


def handle_namelist(race_path, workbook_path):
    # sheet
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb['Sheet1']

    # 清理表格
    clean_up(sheet, 6, 3)

    #
    row = 6
    for dir_name in os.listdir(race_path):
        res = dir_name.split(" ")
        # 序号
        fill_cell(sheet.cell(row, 1), res[0])
        # 团队名
        fill_cell(sheet.cell(row, 2), res[1])

        team_dir = race_path + "\\" + dir_name
        files = os.listdir(team_dir)
        cell = sheet.cell(row, column=3)

        if len(files) == 1:
            fill_cell_link(cell, files[0], team_dir + "\\" + files[0])
        else:
            compress_file = None
            dir_path = None
            for file in files:
                if os.path.splitext(file)[-1][1:] == "zip" or os.path.splitext(file)[-1][1:] == "rar":
                    compress_file = file
                elif os.path.isdir(team_dir + "\\" + file):
                    dir_path = file

            # 压缩文件
            if compress_file is not None and dir_path is not None:
                work_dir = team_dir + "\\" + dir_path
                star_file = None
                word_file = None  # contains wps doc docx
                pdf_file = None
                ppt_file = None  # ppt pptx

                for file in os.listdir(work_dir):
                    tail = os.path.splitext(file)[-1][1:]
                    if file.__contains__("【@】"):
                        star_file = file
                    elif tail == "docx" or tail == "doc" or tail == "wps":
                        word_file = file
                    elif tail == "pdf":
                        pdf_file = file
                    elif tail == "pptx" or tail == "ppt":
                        ppt_file = file
                if star_file is not None:
                    fill_cell_link(cell, compress_file, work_dir + "\\" + star_file)
                elif word_file is not None:
                    fill_cell_link(cell, compress_file, work_dir + "\\" + word_file)
                elif pdf_file is not None:
                    fill_cell_link(cell, compress_file, work_dir + "\\" + pdf_file)
                elif ppt_file is not None:
                    fill_cell_link(cell, compress_file, work_dir + "\\" + ppt_file)
                else:
                    fill_cell_link(cell, compress_file, work_dir)
            else:
                fill_cell_link(cell, files[0], team_dir)

        # 累加和
        cell = sheet.cell(row, column=COLUMN_SUM)
        cell.value = "=SUM({}:{})".format("D" + str(row), "K" + str(row))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        # 设置列高
        sheet.row_dimensions[row].height = 100

        row += 1

    wb.save(workbook_path)


def fill_cell_link(cell, file_name, link):
    cell.value = file_name
    cell.hyperlink = link
    font = Font('宋体', color='0563C1', bold=False, size=12)
    cell.font = font
    # cell.style = "Hyperlink"
    cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)


def fill_cell(cell, content):
    cell.value = content
    cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
    font = Font('宋体', color='000000', bold=False, size=FONT_SIZE)
    cell.font = font


def alignment():
    return Alignment(horizontal='left', vertical='center', wrapText=True)


def is_empty(value) -> bool:
    return value is None or value == ''


def clean_up(sheet: Worksheet, row_start, column_max):
    _row_start = row_start
    cell = sheet.cell(_row_start, 1)
    while not is_empty(cell.value):
        for i in range(1, column_max + 1):
            cell = sheet.cell(_row_start, i)
            cell.value = None
        sheet.cell(_row_start, COLUMN_SUM).value = None
        _row_start += 1
        cell = sheet.cell(_row_start, 1)
